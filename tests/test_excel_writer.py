"""Tests for Excel writer module."""

from pathlib import Path

from openpyxl import load_workbook

from invoice_scrapper.config import DEFAULT_FIELDS
from invoice_scrapper.excel_writer import write_excel
from invoice_scrapper.models import InvoiceData, LineItem


def test_correct_row_count(tmp_path: Path):
    invoices = [
        InvoiceData(
            ficheiro="a.pdf",
            fornecedor="A",
            line_items=[LineItem(), LineItem()],
        ),
        InvoiceData(
            ficheiro="b.pdf",
            fornecedor="B",
            line_items=[LineItem()],
        ),
    ]
    out = tmp_path / "out.xlsx"
    write_excel(invoices, out)

    wb = load_workbook(out)
    ws = wb.active
    # 1 header + 3 data rows
    assert ws.max_row == 4


def test_invoice_fields_repeated(tmp_path: Path):
    invoices = [
        InvoiceData(
            ficheiro="test.pdf",
            fornecedor="ACME",
            line_items=[LineItem(quantidade="1"), LineItem(quantidade="2")],
        ),
    ]
    out = tmp_path / "out.xlsx"
    write_excel(invoices, out)

    wb = load_workbook(out)
    ws = wb.active
    # Find fornecedor column
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    forn_col = headers.index("fornecedor") + 1
    assert ws.cell(row=2, column=forn_col).value == "ACME"
    assert ws.cell(row=3, column=forn_col).value == "ACME"


def test_missing_fields_blank(tmp_path: Path):
    invoices = [InvoiceData(ficheiro="empty.pdf")]
    out = tmp_path / "out.xlsx"
    write_excel(invoices, out)

    wb = load_workbook(out)
    ws = wb.active
    # All data cells except ficheiro should be empty or blank
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    fich_col = headers.index("ficheiro") + 1
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=c).value
        if c == fich_col:
            assert val == "empty.pdf"
        else:
            assert val is None or val == ""


def test_column_headers_match_fields(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    write_excel([], out)

    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    expected = [*DEFAULT_FIELDS, "ficheiro"]
    assert headers == expected


def test_custom_fields(tmp_path: Path):
    custom = ["campo1", "campo2"]
    invoices = [InvoiceData(ficheiro="x.pdf")]
    out = tmp_path / "out.xlsx"
    write_excel(invoices, out, fields=custom)

    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["campo1", "campo2", "ficheiro"]


def test_creates_parent_dirs(tmp_path: Path):
    out = tmp_path / "nested" / "dir" / "out.xlsx"
    write_excel([], out)
    assert out.exists()


def test_empty_invoices_list(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    write_excel([], out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1  # Header only
