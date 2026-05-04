"""Excel output generation for invoice data."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from invoice_scrapper.config import DEFAULT_FIELDS
from invoice_scrapper.models import InvoiceData

# Column headers in display order
COLUMNS = [
    *DEFAULT_FIELDS,
    "ficheiro",
]


def write_excel(
    invoices: list[InvoiceData],
    output_path: Path,
    fields: list[str] | None = None,
) -> Path:
    """Write invoice data to an Excel file.

    Args:
        invoices: List of extracted invoice data.
        output_path: Path for the output .xlsx file.
        fields: Custom field list for column headers.

    Returns:
        Path to the generated file.
    """
    columns = [*(fields or DEFAULT_FIELDS), "ficheiro"]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas"

    # Write header row
    for col_idx, header in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    row_num = 2
    for invoice in invoices:
        for row_data in invoice.to_rows():
            for col_idx, header in enumerate(columns, 1):
                ws.cell(
                    row=row_num,
                    column=col_idx,
                    value=row_data.get(header, ""),
                )
            row_num += 1

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row in range(2, row_num):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_len + 2, 50
        )

    wb.save(output_path)
    return output_path
